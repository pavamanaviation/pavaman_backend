import io
import json
import os
import random
import shutil
from datetime import datetime, timedelta
from django.utils import timezone
from io import BytesIO
import boto3
import openpyxl
import pytz
from aiohttp import ClientError
from botocore.exceptions import ClientError as BotoClientError
from django.conf import settings
from django.contrib.sessions.models import Session
from django.db import IntegrityError
from django.db.models import (
    Sum, Count, Avg, Q, F
)
from django.db.models.functions import TruncMonth
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook
from .msg91 import send_verify_mobile
from pavaman_backend.models import (
    CustomerRegisterDetails, PavamanAdminDetails, CategoryDetails,
    SubCategoryDetails, ProductsDetails, PaymentDetails, OrderProducts,
    FeedbackRating, CustomerAddress
)
import traceback
import io
@csrf_exempt
def add_admin(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            username = data.get('username')
            email = data.get('email')
            mobile_no = data.get('mobile_no')
            password = data.get('password')
            status = data.get('status', 1)
            if not username or not email or not password:
                return JsonResponse({"error": "Username, email, and password are required.", "status_code": 400}, status=400)
            if PavamanAdminDetails.objects.filter(username=username).exists():
                return JsonResponse({"error": "Username already exists. Please choose a different username.", "status_code": 409}, status=409)
            if PavamanAdminDetails.objects.filter(email=email).exists():
                return JsonResponse({"error": "Email already exists. Please use a different email.", "status_code": 409}, status=409)
            admin = PavamanAdminDetails(username=username, email=email, password=password, status=int(status))
            admin.save()
            return JsonResponse({"message": "Admin added successfully", "id": admin.id, "status_code": 201}, status=201)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON data in the request body.", "status_code": 400}, status=400)
        except IntegrityError:
            return JsonResponse({"error": "Database integrity error.", "status_code": 500}, status=500)
        except Exception as e:
            return JsonResponse({"error": f"An unexpected error occurred: {str(e)}", "status_code": 500}, status=500)
    return JsonResponse({"error": "Invalid HTTP method. Only POST is allowed.", "status_code": 405}, status=405)
# @csrf_exempt
# def admin_login(request):
#     if request.method == "POST":
#         try:
#             data = json.loads(request.body)
#             email = data.get('email', '').strip().lower()
#             password = data.get('password', '')
#             if not email or not password:
#                 return JsonResponse({"error": "Email and password are required.", "status_code": 400}, status=400)
#             admin = PavamanAdminDetails.objects.filter(email=email).first()
#             if not admin:
#                 return JsonResponse({"error": "Email not found.", "status_code": 404}, status=404)
#             if admin.password != password:
#                 return JsonResponse({"error": "Invalid email or password.", "status_code": 401}, status=401)
#             if admin.status != 1:
#                 return JsonResponse({"error": "Your account is inactive. Contact support.", "status_code": 403}, status=403)
#             otp = random.randint(100000, 999999)
#             admin.otp = otp
#             admin.save()
#             success = send_otp_sms(admin.mobile_no, otp)
#             if not success:
#                 return JsonResponse({"error": "Failed to send OTP. Try again later.", "status_code": 500}, status=500)
#             return JsonResponse({
#                 "message": "OTP sent to your registered mobile number.",
#                 "status_code": 200,
#                 "email": admin.email
#             }, status=200)
#         except json.JSONDecodeError:
#             return JsonResponse({"error": "Invalid JSON data in the request body.", "status_code": 400}, status=400)
#         except Exception as e:
#             return JsonResponse({"error": f"An unexpected error occurred: {str(e)}", "status_code": 500}, status=500)
#     return JsonResponse({"error": "Invalid HTTP method. Only POST is allowed.", "status_code": 405}, status=405)

@csrf_exempt
def admin_login(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            email = data.get('email', '').strip().lower()
            password = data.get('password', '')
            if not email or not password:
                return JsonResponse({"error": "Email and password are required.", "status_code": 400}, status=400)
            admin = PavamanAdminDetails.objects.filter(email=email).first()
            if not admin:
                return JsonResponse({"error": "Email not found.", "status_code": 404}, status=404)
            if admin.password != password:
                return JsonResponse({"error": "Invalid email or password.", "status_code": 401}, status=401)
            if admin.status != 1:
                return JsonResponse({"error": "Your account is inactive. Contact support.", "status_code": 403}, status=403)
            otp = random.randint(100000, 999999)
            admin.otp = otp
            admin.save()
            success = send_otp_sms(admin.mobile_no, otp)
            if not success:
                return JsonResponse({"error": "Failed to send OTP. Try again later.", "status_code": 500}, status=500)
            return JsonResponse({
                "message": "OTP sent to your registered mobile number.",
                "status_code": 200,
                "email": admin.email
            }, status=200)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON data in the request body.", "status_code": 400}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"An unexpected error occurred: {str(e)}", "status_code": 500}, status=500)
    return JsonResponse({"error": "Invalid HTTP method. Only POST is allowed.", "status_code": 405}, status=405)

# @csrf_exempt
# def admin_verify_otp(request):
#     if request.method == "POST":
#         try:
#             data = json.loads(request.body)
#             email = data.get("email")
#             otp = data.get("otp")
#             if not email or not otp:
#                 return JsonResponse({"error": "Email and OTP are required.", "status_code": 400}, status=400)
#             admin = PavamanAdminDetails.objects.filter(email=email).first()
#             if not admin:
#                 return JsonResponse({"error": "Invalid email.", "status_code": 404}, status=404)
#             if str(admin.otp) != str(otp):
#                 return JsonResponse({"error": "Invalid OTP.", "status_code": 401}, status=401)
#             admin.otp = None
#             admin.save()
#             request.session['admin_id'] = admin.id
#             request.session['admin_email'] = admin.email
#             request.session['admin_username'] = admin.username
#             request.session.modified = True
#             return JsonResponse({"message": "OTP verified.Login successful.", "username": admin.username, "email": admin.email, "id": admin.id, "status_code": 200}, status=200)
#         except json.JSONDecodeError:
#             return JsonResponse({"error": "Invalid JSON data.", "status_code": 400}, status=400)
#         except Exception as e:
#             return JsonResponse({"error": f"Unexpected error: {str(e)}", "status_code": 500}, status=500)
#     return JsonResponse({"error": "Only POST method allowed.", "status_code": 405}, status=405)

@csrf_exempt
def admin_verify_otp(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            email = data.get("email").strip().lower()
            otp = data.get("otp")
            if not email or not otp:
                return JsonResponse({"error": "Email and OTP are required.", "status_code": 400}, status=400)
            admin = PavamanAdminDetails.objects.filter(email=email).first()
            if not admin:
                return JsonResponse({"error": "Invalid email.", "status_code": 404}, status=404)
            if str(admin.otp) != str(otp):
                return JsonResponse({"error": "Invalid OTP.", "status_code": 401}, status=401)
            admin.otp = None
            admin.save()
            request.session['admin_id'] = admin.id
            request.session['admin_email'] = admin.email
            request.session['admin_username'] = admin.username
            request.session.modified = True
            return JsonResponse({"message": "OTP verified.Login successful.", "username": admin.username, "email": admin.email, "id": admin.id, "status_code": 200}, status=200)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON data.", "status_code": 400}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Unexpected error: {str(e)}", "status_code": 500}, status=500)
    return JsonResponse({"error": "Only POST method allowed.", "status_code": 405}, status=405)

def send_otp_sms(mobile_no, otp):
    try:
        send_verify_mobile([mobile_no], otp)
        return True
    except Exception as e:
        print(f"Failed to send OTP to {mobile_no}: {e}")
        return False
@csrf_exempt
def admin_logout(request):
    if request.method == "POST":
        try:
            if 'admin_id' in request.session:
                request.session.flush()
                return JsonResponse({"message": "Logout successful.", "status_code": 200}, status=200)
            else:
                return JsonResponse({"error": "No active session found.", "status_code": 400}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"An unexpected error occurred: {str(e)}", "status_code": 500}, status=500)
    return JsonResponse({"error": "Invalid HTTP method. Only POST is allowed.", "status_code": 405}, status=405)
@csrf_exempt
def add_category(request):
    if request.method == 'POST':
        try:
            data = request.POST
            category_name = data.get('category_name')
            admin_id = data.get('admin_id')
            category_status = 1
            if not admin_id:
                return JsonResponse({"error": "Admin is not logged in.", "status_code": 401}, status=401)
            try:
                admin_data = PavamanAdminDetails.objects.get(id=admin_id)
            except PavamanAdminDetails.DoesNotExist:
                return JsonResponse({"error": "Admin session expired or invalid.", "status_code": 401}, status=401)
            if CategoryDetails.objects.filter(category_name=category_name).exists():
                return JsonResponse({"error": "Category name already exists.", "status_code": 409}, status=409)
            if 'category_image' not in request.FILES:
                return JsonResponse({"error": "Category image file is required.", "status_code": 400}, status=400)
            category_image = request.FILES['category_image']
            allowed_extensions = ['png', 'jpg', 'jpeg']
            file_name, file_extension = os.path.splitext(category_image.name)
            file_extension = file_extension.lower().lstrip('.')
            if file_extension not in allowed_extensions:
                return JsonResponse({
                    "error": f"Invalid file type. Allowed types: {', '.join(allowed_extensions)}",
                    "status_code": 400
                }, status=400)
            safe_category_name = category_name.replace(' ', '_').replace('/', '_')
            safe_file_name = file_name.replace(' ', '_').replace('/', '_')
            s3_file_key = f"static/images/category/{safe_category_name}_{safe_file_name}.{file_extension}"
            s3 = boto3.client(
                's3',
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                region_name=settings.AWS_S3_REGION_NAME
            )
            s3.upload_fileobj(
                category_image,
                settings.AWS_STORAGE_BUCKET_NAME,
                s3_file_key,
                ExtraArgs={'ContentType': category_image.content_type}
            )
            image_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{s3_file_key}"
            current_time = datetime.utcnow() + timedelta(hours=5, minutes=30)
            category = CategoryDetails(
                category_name=category_name,
                admin=admin_data,
                category_image=s3_file_key,
                category_status=category_status,
                created_at=current_time
            )
            category.save()
            return JsonResponse({
                "message": "Category added successfully",
                "category_id": category.id,
                "category_image_url": image_url,
                "category_status": category.category_status,
                "status_code": 201
            }, status=201)
        except Exception as e:
            return JsonResponse({
                "error": f"An unexpected error occurred: {str(e)}",
                "status_code": 500
            }, status=500)
    return JsonResponse({"error": "Invalid HTTP method. Only POST is allowed.", "status_code": 405}, status=405)
@csrf_exempt
def view_categories(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            admin_id = data.get('admin_id')
            if not admin_id:
                return JsonResponse({"error": "Admin Id is required.", "status_code": 400}, status=400)
            admin_data = PavamanAdminDetails.objects.filter(id=admin_id).first()
            if not admin_data:
                return JsonResponse({"error": "Admin not found or session expired.", "status_code": 401}, status=401)
            categories = CategoryDetails.objects.filter(admin_id=admin_id, category_status=1)
            if not categories.exists():
                return JsonResponse({"message": "No category details found", "status_code": 200}, status=200)
            category_list = []
            for category in categories:
                image_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{category.category_image}"
                category_list.append({
                    "category_id": str(category.id),
                    "category_name": category.category_name,
                    "category_image_url": image_url
                })
            return JsonResponse(
                {"message": "Categories retrieved successfully.", "categories": category_list, "status_code": 200},
                status=200
            )
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON data.", "status_code": 400}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"An unexpected error occurred: {str(e)}", "status_code": 500}, status=500)
    return JsonResponse({"error": "Invalid HTTP method. Only POST is allowed.", "status_code": 405}, status=405)
# @csrf_exempt
# def edit_category(request):
#     if request.method == 'POST':
#         try:
#             data = request.POST
#             category_id = data.get('category_id')
#             category_name = data.get('category_name').lower()
#             admin_id = data.get('admin_id')
#             if not admin_id:
#                 return JsonResponse({"error": "Admin is not logged in.", "status_code": 401}, status=401)
#             if not category_id:
#                 return JsonResponse({"error": "Category ID is required.", "status_code": 400}, status=400)
#             admin_data = PavamanAdminDetails.objects.filter(id=admin_id).first()
#             if not admin_data:
#                 return JsonResponse({"error": "Admin not found or session expired.", "status_code": 401}, status=401)
#             category = CategoryDetails.objects.filter(id=category_id, admin=admin_data, category_status=1).first()
#             if not category:
#                 return JsonResponse({"error": "Category not found.", "status_code": 404}, status=404)
#             if CategoryDetails.objects.filter(category_name=category_name).exclude(id=category_id).exists():
#                 return JsonResponse({"error": "Category name already exists.", "status_code": 409}, status=409)
#             category.category_name = category_name
#             formatted_category_name = category_name.replace(' ', '_').replace('/', '_')
#             s3 = boto3.client(
#                 's3',
#                 aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
#                 aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
#                 region_name=settings.AWS_S3_REGION_NAME
#             )
#             if 'category_image' in request.FILES:
#                 category_image = request.FILES['category_image']
#                 allowed_extensions = ['png', 'jpg', 'jpeg']
#                 file_extension = category_image.name.split('.')[-1].lower()
#                 if file_extension not in allowed_extensions:
#                     return JsonResponse({
#                         "error": f"Invalid file type. Allowed types: {', '.join(allowed_extensions)}",
#                         "status_code": 400
#                     }, status=400)
#                 image_name = f"{formatted_category_name}_{category_image.name}"
#                 s3_file_key = f"static/images/category/{image_name}"
#                 if category.category_image:
#                     try:
#                         s3.delete_object(Bucket=settings.AWS_STORAGE_BUCKET_NAME, Key=category.category_image)
#                     except Exception as delete_err:
#                         print("Warning: Could not delete old image:", delete_err)
#                 s3.upload_fileobj(
#                     category_image,
#                     settings.AWS_STORAGE_BUCKET_NAME,
#                     s3_file_key,
#                     ExtraArgs={'ContentType': category_image.content_type}
#                 )
#                 category.category_image = s3_file_key
#             elif category.category_image:
#                 old_key = category.category_image
#                 original_file_name = old_key.split('/')[-1].split('_')[-1] 
#                 new_key = f"static/images/category/{formatted_category_name}_{original_file_name}"
#                 if old_key != new_key:
#                     try:
#                         s3.copy_object(
#                             Bucket=settings.AWS_STORAGE_BUCKET_NAME,
#                             CopySource={'Bucket': settings.AWS_STORAGE_BUCKET_NAME, 'Key': old_key},
#                             Key=new_key,
#                             MetadataDirective='REPLACE',
#                             ContentType=f'image/{original_file_name.split(".")[-1]}'
#                         )
#                         s3.delete_object(Bucket=settings.AWS_STORAGE_BUCKET_NAME, Key=old_key)
#                         category.category_image = new_key
#                     except Exception as rename_err:
#                         print("Warning: Could not rename image on S3:", rename_err)
#             category.save()
#             image_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{category.category_image}" if category.category_image else None
#             return JsonResponse({
#                 "message": "Category updated successfully.",
#                 "category_id": str(category.id),
#                 "category_name": category.category_name,
#                 "category_image_url": image_url,
#                 "status_code": 200
#             }, status=200)
#         except Exception as e:
#             return JsonResponse({"error": f"An unexpected error occurred: {str(e)}", "status_code": 500}, status=500)
#     return JsonResponse({"error": "Invalid HTTP method. Only POST is allowed.", "status_code": 405}, status=405)
# @csrf_exempt
# def delete_category(request):
#     if request.method == 'POST':
#         try:
#             try:
#                 data = json.loads(request.body)
#             except json.JSONDecodeError:
#                 return JsonResponse({"error": "Invalid JSON data.", "status_code": 400}, status=400)
#             category_id = data.get('category_id')
#             admin_id = data.get('admin_id')
#             print(f"Admin ID: {admin_id}, Category ID: {category_id}")
#             if not admin_id:
#                 return JsonResponse({"error": "Admin is not logged in.", "status_code": 401}, status=401)
#             if not category_id:
#                 return JsonResponse({"error": "Category ID is required.", "status_code": 400}, status=400)
#             admin_data = PavamanAdminDetails.objects.filter(id=admin_id).first()
#             if not admin_data:
#                 return JsonResponse({"error": "Admin not found or session expired.", "status_code": 401}, status=401)
#             category = CategoryDetails.objects.filter(id=category_id, admin=admin_data).first()
#             if not category:
#                 return JsonResponse({"error": "Category not found or you do not have permission to delete this category.", "status_code": 404}, status=404)
#             if category.category_image:
#                 s3 = boto3.client(
#                     's3',
#                     aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
#                     aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
#                     region_name=settings.AWS_S3_REGION_NAME
#                 )
#                 try:
#                     s3.delete_object(
#                         Bucket=settings.AWS_STORAGE_BUCKET_NAME,
#                         Key=category.category_image
#                     )
#                 except ClientError as e:
#                     print(f"S3 deletion error: {e}")
#             category.delete()
#             return JsonResponse({
#                 "message": "Category deleted successfully.",
#                 "status_code": 200
#             }, status=200)
#         except Exception as e:
#             return JsonResponse({"error": f"An unexpected error occurred: {str(e)}", "status_code": 500}, status=500)
#     return JsonResponse({"error": "Invalid HTTP method. Only POST is allowed.", "status_code": 405}, status=405)

@csrf_exempt
def edit_category(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body.decode('utf-8'))
            category_id = data.get('category_id')
            category_name = data.get('category_name')
            admin_id = request.session.get('admin_id')
            if not admin_id:
                return JsonResponse({"error": "Admin is not logged in.", "status_code": 401}, status=401)
            if not category_id:
                return JsonResponse({"error": "Category ID is required.", "status_code": 400}, status=400)
            admin_data = PavamanAdminDetails.objects.filter(id=admin_id).first()
            if not admin_data:
                return JsonResponse({"error": "Admin not found or session expired.", "status_code": 401}, status=401)
            category = CategoryDetails.objects.filter(id=category_id, admin=admin_data, category_status=1).first()
            if not category:
                return JsonResponse({"error": "Category not found.", "status_code": 404}, status=404)
            if CategoryDetails.objects.filter(category_name=category_name).exclude(id=category_id).exists():
                return JsonResponse({"error": "Category name already exists.", "status_code": 409}, status=409)
            category.category_name = category_name
            formatted_category_name = category_name.replace(' ', '').replace('/', '')
            s3 = boto3.client(
                's3',
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                region_name=settings.AWS_S3_REGION_NAME
            )
            if 'category_image' in request.FILES:
                category_image = request.FILES['category_image']
                allowed_extensions = ['png', 'jpg', 'jpeg']
                file_extension = category_image.name.split('.')[-1].lower()
                if file_extension not in allowed_extensions:
                    return JsonResponse({
                        "error": f"Invalid file type. Allowed types: {', '.join(allowed_extensions)}",
                        "status_code": 400
                    }, status=400)
                image_name = f"{formatted_category_name}_{category_image.name}"
                s3_file_key = f"static/images/category/{image_name}"
                if category.category_image:
                    try:
                        s3.delete_object(Bucket=settings.AWS_STORAGE_BUCKET_NAME, Key=category.category_image)
                    except Exception as delete_err:
                        print("Warning: Could not delete old image:", delete_err)
                s3.upload_fileobj(
                    category_image,
                    settings.AWS_STORAGE_BUCKET_NAME,
                    s3_file_key,
                    ExtraArgs={'ContentType': category_image.content_type}
                )
                category.category_image = s3_file_key
            elif category.category_image:
                old_key = category.category_image
                original_file_name = old_key.split('/')[-1].split('_')[-1]
                new_key = f"static/images/category/{formatted_category_name}_{original_file_name}"
                if old_key != new_key:
                    try:
                        s3.copy_object(
                            Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                            CopySource={'Bucket': settings.AWS_STORAGE_BUCKET_NAME, 'Key': old_key},
                            Key=new_key,
                            MetadataDirective='REPLACE',
                            ContentType=f'image/{original_file_name.split(".")[-1]}'
                        )
                        s3.delete_object(Bucket=settings.AWS_STORAGE_BUCKET_NAME, Key=old_key)
                        category.category_image = new_key
                    except Exception as rename_err:
                        print("Warning: Could not rename image on S3:", rename_err)
            category.save()
            image_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{category.category_image}" if category.category_image else None
            return JsonResponse({
                "message": "Category updated successfully.",
                "category_id": str(category.id),
                "category_name": category.category_name,
                "category_image_url": image_url,
                "status_code": 200
            }, status=200)
        except Exception as e:
            return JsonResponse({"error": f"An unexpected error occurred: {str(e)}", "status_code": 500}, status=500)
    return JsonResponse({"error": "Invalid HTTP method. Only POST is allowed.", "status_code": 405}, status=405)

@csrf_exempt
def delete_category(request):
    if request.method == 'POST':
        try:
            try:
                data = json.loads(request.body)
            except json.JSONDecodeError:
                return JsonResponse({"error": "Invalid JSON data.", "status_code": 400}, status=400)
            category_id = data.get('category_id')
            admin_id = data.get('admin_id')
            print(f"Admin ID: {admin_id}, Category ID: {category_id}")
            if not admin_id:
                return JsonResponse({"error": "Admin is not logged in.", "status_code": 401}, status=401)
            if not category_id:
                return JsonResponse({"error": "Category ID is required.", "status_code": 400}, status=400)
            admin_data = PavamanAdminDetails.objects.filter(id=admin_id).first()
            if not admin_data:
                return JsonResponse({"error": "Admin not found or session expired.", "status_code": 401}, status=401)
            category = CategoryDetails.objects.filter(id=category_id, admin=admin_data).first()
            if not category:
                return JsonResponse({"error": "Category not found or you do not have permission to delete this category.", "status_code": 404}, status=404)
            if category.category_image:
                s3 = boto3.client(
                    's3',
                    aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                    aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                    region_name=settings.AWS_S3_REGION_NAME
                )
                try:
                    s3.delete_object(
                        Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                        Key=category.category_image
                    )
                except ClientError as e:
                    print(f"S3 deletion error: {e}")
            category.delete()
            return JsonResponse({
                "message": "Category deleted successfully.",
                "status_code": 200
            }, status=200)
        except Exception as e:
            return JsonResponse({"error": f"An unexpected error occurred: {str(e)}", "status_code": 500}, status=500)
    return JsonResponse({"error": "Invalid HTTP method. Only POST is allowed.", "status_code": 405}, status=405)

@csrf_exempt
def add_subcategory(request):
    if request.method == 'POST':
        try:
            data = request.POST
            sub_category_name = data.get('sub_category_name') 
            category_id = data.get('category_id')
            admin_id = data.get('admin_id')
            sub_category_status = 1
            if not sub_category_name:
                return JsonResponse({"error": "Subcategory name is required.", "status_code": 400}, status=400)
            if not admin_id:
                return JsonResponse({"error": "Admin is not logged in.", "status_code": 401}, status=401)
            sub_category_name = sub_category_name.lower()
            try:
                admin_data = PavamanAdminDetails.objects.get(id=admin_id)
            except PavamanAdminDetails.DoesNotExist:
                return JsonResponse({"error": "Admin session expired or invalid.", "status_code": 401}, status=401)
            try:
                category = CategoryDetails.objects.get(id=category_id, admin=admin_data)
            except CategoryDetails.DoesNotExist:
                return JsonResponse({"error": "Category not found.", "status_code": 404}, status=404)
            if SubCategoryDetails.objects.filter(sub_category_name=sub_category_name, category=category).exists():
                return JsonResponse({
                    "error": f"Subcategory '{sub_category_name}' already exists under category '{category.category_name}'.",
                    "status_code": 409
                }, status=409)
            existing_subcategory = SubCategoryDetails.objects.filter(sub_category_name=sub_category_name).exclude(category=category).first()
            if existing_subcategory:
                return JsonResponse({
                    "error": f"Subcategory '{sub_category_name}' already exists under a different category '{existing_subcategory.category.category_name}'.",
                    "status_code": 409
                }, status=409)
            sub_category_image = request.FILES.get('sub_category_image', None)
            if not sub_category_image:
                return JsonResponse({"error": "Subcategory image file is required.", "status_code": 400}, status=400)
            allowed_extensions = ['png', 'jpg', 'jpeg']
            file_name, file_extension = os.path.splitext(sub_category_image.name)
            file_extension = file_extension.lower().lstrip('.') 
            if file_extension not in allowed_extensions:
                return JsonResponse({
                    "error": f"Invalid file type. Allowed types: {', '.join(allowed_extensions)}",
                    "status_code": 400
                }, status=400)
            safe_subcat_name = sub_category_name.replace(' ', '_').replace('/', '_')
            safe_file_name = file_name.replace(' ', '_').replace('/', '_')
            s3_file_key = f"static/images/subcategory/{safe_subcat_name}_{safe_file_name}.{file_extension}"
            s3 = boto3.client(
                's3',
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                region_name=settings.AWS_S3_REGION_NAME
            )
            s3.upload_fileobj(
                sub_category_image,
                settings.AWS_STORAGE_BUCKET_NAME,
                s3_file_key,
                ExtraArgs={'ContentType': sub_category_image.content_type}
            )
            image_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{s3_file_key}"
            current_time = datetime.utcnow() + timedelta(hours=5, minutes=30)
            subcategory = SubCategoryDetails(
                sub_category_name=sub_category_name,
                category=category,
                sub_category_image=s3_file_key,
                sub_category_status=sub_category_status,
                admin=admin_data,
                created_at=current_time
            )
            subcategory.save()
            return JsonResponse({
                "message": "Subcategory added successfully",
                "subcategory_id": subcategory.id,
                "category_id": category.id,
                "category_name": category.category_name,
                "subcategory_image_url": image_url,
                "subcategory_status": subcategory.sub_category_status,
                "status_code": 201
            }, status=201)
        except Exception as e:
            return JsonResponse({"error": f"An unexpected error occurred: {str(e)}", "status_code": 500}, status=500)
    return JsonResponse({"error": "Invalid HTTP method. Only POST is allowed.", "status_code": 405}, status=405)
@csrf_exempt
def view_subcategories(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body.decode('utf-8'))
            admin_id = data.get('admin_id')
            category_id = data.get('category_id')
            if not admin_id or not category_id:
                return JsonResponse({
                    "error": "Admin id and Category id are required.",
                    "status_code": 400
                }, status=400)
            try:
                admin = PavamanAdminDetails.objects.get(id=admin_id)
                category = CategoryDetails.objects.get(id=category_id, admin=admin)
            except PavamanAdminDetails.DoesNotExist:
                return JsonResponse({"error": "Admin not found or session expired.", "status_code": 404}, status=404)
            except CategoryDetails.DoesNotExist:
                return JsonResponse({"error": "Category not found.", "status_code": 404}, status=404)
            subcategories = SubCategoryDetails.objects.filter(category=category)
            subcategory_list = []
            for subcat in subcategories:
                if subcat.sub_category_image:
                    image_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.amazonaws.com/{subcat.sub_category_image}"
                else:
                    image_url = None
                subcategory_list.append({
                    'id': subcat.id,
                    'sub_category_name': subcat.sub_category_name,
                    'sub_category_image': image_url,
                })
            if not subcategory_list:
                return JsonResponse({
                    "message": "No subcategories found.",
                    "status_code": 200,
                    "subcategories": []
                }, status=200)
            return JsonResponse({
                "message": "Subcategories retrieved successfully.",
                "status_code": 200,
                "category_id": category.id,
                "category_name": category.category_name,
                "subcategories": subcategory_list
            }, status=200)
        except json.JSONDecodeError:
            return JsonResponse({
                "error": "Invalid JSON format.",
                "status_code": 400
            }, status=400)
        except Exception as e:
            return JsonResponse({
                "error": f"An unexpected error occurred: {str(e)}",
                "status_code": 500
            }, status=500)

    return JsonResponse({"error": "Invalid HTTP method. Only POST is allowed.", "status_code": 405}, status=405)
@csrf_exempt
def edit_subcategory(request):
    if request.method == 'POST':
        try:
            data = request.POST
            subcategory_id = data.get('subcategory_id')
            sub_category_name = data.get('subcategory_name').lower()
            category_id = data.get('category_id')
            admin_id = data.get('admin_id')
            if not admin_id:
                return JsonResponse({"error": "Admin is not logged in.", "status_code": 401}, status=401)
            if not subcategory_id:
                return JsonResponse({"error": "Subcategory ID is required.", "status_code": 400}, status=400)
            if not sub_category_name:
                return JsonResponse({"error": "Subcategory Name is required.", "status_code": 400}, status=400)
            admin_data = PavamanAdminDetails.objects.filter(id=admin_id).first()
            if not admin_data:
                return JsonResponse({"error": "Admin not found or session expired.", "status_code": 401}, status=401)
            category = CategoryDetails.objects.filter(id=category_id, admin=admin_data).first()
            if not category:
                return JsonResponse({"error": "Category not found.", "status_code": 404}, status=404)
            subcategory = SubCategoryDetails.objects.filter(id=subcategory_id, category=category).first()
            if not subcategory:
                return JsonResponse({"error": "Subcategory not found.", "status_code": 404}, status=404)
            existing_subcategory = SubCategoryDetails.objects.filter(
                sub_category_name=sub_category_name, category=category
            ).exclude(id=subcategory_id).first()
            if existing_subcategory:
                return JsonResponse({
                    "error": f"Subcategory name already exists under {category.category_name}",
                    "status_code": 409
                }, status=409)
            subcategory.sub_category_name = sub_category_name
            formatted_sub_name = sub_category_name.replace(' ', '_').replace('/', '_')
            s3 = boto3.client(
                's3',
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                region_name=settings.AWS_S3_REGION_NAME
            )
            if 'sub_category_image' in request.FILES:
                subcategory_image = request.FILES['sub_category_image']
                allowed_extensions = ['png', 'jpg', 'jpeg']
                file_extension = subcategory_image.name.split('.')[-1].lower()
                if file_extension not in allowed_extensions:
                    return JsonResponse({"error": f"Invalid file type. Allowed types: {', '.join(allowed_extensions)}", "status_code": 400}, status=400)
                image_name = f"{formatted_sub_name}_{subcategory_image.name}"
                s3_file_key = f"static/images/subcategory/{image_name}"
                if subcategory.sub_category_image:
                    try:
                        s3.delete_object(Bucket=settings.AWS_STORAGE_BUCKET_NAME, Key=subcategory.sub_category_image)
                    except Exception as e:
                        print("Warning: Failed to delete old subcategory image:", e)
                s3.upload_fileobj(
                    subcategory_image,
                    settings.AWS_STORAGE_BUCKET_NAME,
                    s3_file_key,
                    ExtraArgs={'ContentType': subcategory_image.content_type}
                )
                subcategory.sub_category_image = s3_file_key
            elif subcategory.sub_category_image:
                old_key = subcategory.sub_category_image
                original_file_name = old_key.split('/')[-1].split('_')[-1]
                new_key = f"static/images/subcategory/{formatted_sub_name}_{original_file_name}"
                if old_key != new_key:
                    try:
                        s3.copy_object(
                            Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                            CopySource={'Bucket': settings.AWS_STORAGE_BUCKET_NAME, 'Key': old_key},
                            Key=new_key,
                            MetadataDirective='REPLACE',
                            ContentType=f"image/{original_file_name.split('.')[-1]}"
                        )
                        s3.delete_object(Bucket=settings.AWS_STORAGE_BUCKET_NAME, Key=old_key)
                        subcategory.sub_category_image = new_key
                    except Exception as e:
                        print("Warning: Failed to rename subcategory image:", e)
            subcategory.save()
            image_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{subcategory.sub_category_image}" if subcategory.sub_category_image else None
            return JsonResponse({
                "message": "Subcategory updated successfully.",
                "category_id": subcategory.category.id,
                "category_name": subcategory.category.category_name,
                "subcategory_id": subcategory.id,
                "subcategory_name": subcategory.sub_category_name,
                "subcategory_image_url": image_url,
                "status_code": 200
            }, status=200)
        except Exception as e:
            return JsonResponse({"error": f"An unexpected error occurred: {str(e)}", "status_code": 500}, status=500)
    return JsonResponse({"error": "Invalid HTTP method. Only POST is allowed.", "status_code": 405}, status=405)
@csrf_exempt
def delete_subcategory(request):
    if request.method == 'POST':
        try:
            try:
                data = json.loads(request.body)
            except json.JSONDecodeError:
                return JsonResponse({"error": "Invalid JSON data.", "status_code": 400}, status=400)
            subcategory_id = data.get('subcategory_id')
            category_id = data.get('category_id')
            admin_id = data.get('admin_id')
            print(f"Admin ID: {admin_id}, Category ID: {category_id}, Subcategory ID: {subcategory_id}")
            if not admin_id:
                return JsonResponse({"error": "Admin is not logged in.", "status_code": 401}, status=401)
            if not category_id:
                return JsonResponse({"error": "Category ID is required.", "status_code": 400}, status=400)
            if not subcategory_id:
                return JsonResponse({"error": "Subcategory ID is required.", "status_code": 400}, status=400)
            admin_data = PavamanAdminDetails.objects.filter(id=admin_id).first()
            if not admin_data:
                return JsonResponse({"error": "Admin not found or session expired.", "status_code": 401}, status=401)
            category = CategoryDetails.objects.filter(id=category_id, admin=admin_data).first()
            if not category:
                return JsonResponse({"error": "Category not found under this admin.", "status_code": 404}, status=404)
            subcategory = SubCategoryDetails.objects.filter(id=subcategory_id, category=category).first()
            if not subcategory:
                return JsonResponse({"error": "Subcategory not found.", "status_code": 404}, status=404)
            if subcategory.sub_category_image:
                s3 = boto3.client(
                    's3',
                    aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                    aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                    region_name=settings.AWS_S3_REGION_NAME
                )
                try:
                    s3.delete_object(
                        Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                        Key=subcategory.sub_category_image
                    )
                except ClientError as e:
                    print(f"S3 deletion error: {e}")
            subcategory.delete()
            return JsonResponse({
                "message": "Subcategory deleted successfully.",
                "status_code": 200
            }, status=200)
        except Exception as e:
            return JsonResponse({"error": f"An unexpected error occurred: {str(e)}", "status_code": 500}, status=500)
    return JsonResponse({"error": "Invalid HTTP method. Only POST is allowed.", "status_code": 405}, status=405)
@csrf_exempt
def add_product(request):
    if request.method == 'POST':
        try:
            if request.content_type == "application/json":
                try:
                    data = json.loads(request.body.decode('utf-8'))
                except json.JSONDecodeError:
                    return JsonResponse({"error": "Invalid JSON format.", "status_code": 400}, status=400)
            else:
                data = request.POST.dict()
            product_name = data.get('product_name').lower()
            sku_number = data.get('sku_number')
            hsn_code=data.get('hsn_code')
            price = data.get('price')
            quantity = data.get('quantity')
            discount = data.get('discount') or 0
            gst = data.get('gst')
            description = data.get('description')
            admin_id = data.get('admin_id')
            category_id = data.get('category_id')
            sub_category_id = data.get('sub_category_id')
            if not all([product_name, sku_number,hsn_code, price, quantity, description, admin_id, category_id, sub_category_id]):
                return JsonResponse({"error": "Missing required fields.", "status_code": 400}, status=400)
            try:
                price = float(price)
                quantity = int(quantity)
                discount = int(discount)
                gst = float(gst)
                if discount > price:
                    return JsonResponse({"error": "Discount amount cannot be more than the price.", "status_code": 400}, status=400)
            except ValueError:
                return JsonResponse({"error": "Invalid format for price, quantity, or discount.", "status_code": 400}, status=400)
            availability = "Out of Stock" if quantity == 0 else "Very Few Products Left" if quantity <= 10 else "In Stock"
            try:
                admin = PavamanAdminDetails.objects.get(id=admin_id)
            except PavamanAdminDetails.DoesNotExist:
                return JsonResponse({"error": "Admin not found.", "status_code": 401}, status=401)
            try:
                category = CategoryDetails.objects.get(id=category_id, admin=admin)
            except CategoryDetails.DoesNotExist:
                return JsonResponse({"error": "Category not found", "status_code": 404}, status=404)
            try:
                sub_category = SubCategoryDetails.objects.get(id=sub_category_id, category=category)
            except SubCategoryDetails.DoesNotExist:
                return JsonResponse({"error": "Sub-category not found.", "status_code": 404}, status=404)
            if ProductsDetails.objects.filter(product_name=product_name).exists():
                return JsonResponse({"error": "Product name already exists.", "status_code": 409}, status=409)
            if ProductsDetails.objects.filter(sku_number=sku_number).exists():
                return JsonResponse({"error": "SKU number already exists.", "status_code": 409}, status=409)
            if 'product_images' not in request.FILES:
                return JsonResponse({"error": "Product images are required.", "status_code": 400}, status=400)
            image_files = request.FILES.getlist('product_images')
            if not image_files:
                return JsonResponse({"error": "At least one product image is required.", "status_code": 400}, status=400)
            s3 = boto3.client(
                's3',
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                region_name=settings.AWS_S3_REGION_NAME
            )
            product_images = []
            allowed_image_extensions = ['png', 'jpg', 'jpeg']
            for image in image_files:
                file_extension = image.name.split('.')[-1].lower()
                if file_extension not in allowed_image_extensions:
                    return JsonResponse({"error": f"Invalid image file type. Allowed types: {', '.join(allowed_image_extensions)}", "status_code": 400}, status=400)
                s3_key = f"static/images/products/{product_name.replace(' ', '_')}/{sku_number}_{image.name}"
                try:
                    s3.upload_fileobj(
                        image,
                        settings.AWS_STORAGE_BUCKET_NAME,
                        s3_key,
                        ExtraArgs={'ContentType': image.content_type}
                    )
                except ClientError as e:
                    return JsonResponse({"error": f"Failed to upload product image: {str(e)}", "status_code": 500}, status=500)
                product_images.append(s3_key)
            if 'material_file' not in request.FILES:
                return JsonResponse({"error": "Material file is required.", "status_code": 400}, status=400)
            material_file = request.FILES['material_file']
            allowed_material_extensions = ['pdf', 'doc']
            file_extension = material_file.name.split('.')[-1].lower()
            if file_extension not in allowed_material_extensions:
                return JsonResponse({"error": f"Invalid material file type. Allowed types: {', '.join(allowed_material_extensions)}", "status_code": 400}, status=400)
            material_key = f"static/materials/{product_name.replace(' ', '_')}.{file_extension}"
            try:
                s3.upload_fileobj(
                    material_file,
                    settings.AWS_STORAGE_BUCKET_NAME,
                    material_key,
                    ExtraArgs={'ContentType': material_file.content_type}
                )
            except ClientError as e:
                return JsonResponse({"error": f"Failed to upload material file: {str(e)}", "status_code": 500}, status=500)
            current_time = datetime.utcnow() + timedelta(hours=5, minutes=30)            
            product = ProductsDetails(
                product_name=product_name,
                sku_number=sku_number,
                hsn_code=hsn_code,
                price=price,
                quantity=quantity,
                discount=discount,
                description=description,
                admin=admin,
                category=category,
                gst=gst,
                sub_category=sub_category,
                product_images=product_images,
                material_file=material_key,
                availability=availability,
                created_at=current_time,
                product_status=1,
                cart_status=False
            )
            product.save()
            return JsonResponse({
                "message": "Product added successfully.",
                "category_id": str(product.category.id),
                "category_name": product.category.category_name,
                "subcategory_id": str(product.sub_category.id),
                "sub_category_name": product.sub_category.sub_category_name,
                "product_id": str(product.id),
                "availability": availability,
                "discount": f"{int(product.discount)}%" if isinstance(product.discount, (int, float)) else product.discount,
                "gst": f"{int(product.gst)}%" if isinstance(product.gst, (int, float)) else product.gst,
                "status_code": 201
            }, status=201)
        except Exception as e:
            return JsonResponse({"error": f"Unexpected error: {str(e)}", "status_code": 500}, status=500)
    return JsonResponse({"error": "Invalid request method. Only POST is allowed.", "status_code": 405}, status=405)
@csrf_exempt
def upload_products_excel(request):
    if request.method != 'POST':
        return JsonResponse({"error": "Invalid request method."}, status=405)
    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        return JsonResponse({"error": "Excel file is required."}, status=400)
    category_id = request.POST.get('category_id')
    sub_category_id = request.POST.get('sub_category_id')
    admin_id = request.POST.get('admin_id')
    if not all([category_id, sub_category_id, admin_id]):
        return JsonResponse({"error": "category_id, sub_category_id, and admin_id are required in POST data."}, status=400)
    try:
        admin = PavamanAdminDetails.objects.get(id=admin_id)
        category = CategoryDetails.objects.get(id=category_id, admin=admin)
        sub_category = SubCategoryDetails.objects.get(id=sub_category_id, category=category)
    except Exception as e:
        return JsonResponse({"error": f"Invalid admin/category/sub-category: {str(e)}"}, status=400)
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    all_image_names = []
    all_material_names = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_data = dict(zip(headers, row))
        image_names = str(row_data.get('image_paths') or '').split(',')
        for img_name in image_names:
            img_name = img_name.strip()
            if img_name:
                all_image_names.append(img_name.lower())
        material_name = str(row_data.get('material_paths') or '').strip()
        if material_name:
            all_material_names.append(material_name.lower())
    duplicates_images = set([x for x in all_image_names if all_image_names.count(x) > 1])
    duplicates_materials = set([x for x in all_material_names if all_material_names.count(x) > 1])
    if duplicates_images or duplicates_materials:
        error_parts = []
        if duplicates_images:
            error_parts.append(f"Duplicate image file names: {', '.join(duplicates_images)}")
        if duplicates_materials:
            error_parts.append(f"Duplicate material file names: {', '.join(duplicates_materials)}")
        return JsonResponse({
            "error": " ; ".join(error_parts) + ". Please ensure file names are unique across all products.",
            "products": product_responses,
        }, status=400)
    uploaded_images = {f.name: f for f in request.FILES.getlist('images[]')}
    uploaded_materials = {f.name: f for f in request.FILES.getlist('materials[]')}
    s3 = boto3.client(
        's3',
        aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
        aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
        region_name=settings.AWS_S3_REGION_NAME
    )
    product_responses = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            row_data = dict(zip(headers, row))
            product_name = row_data.get('product_name')
            if not product_name:
                return JsonResponse({"error": f"Product name missing in row {i}", "products": product_responses}, status=400)
            product_name = product_name.lower()
            if ProductsDetails.objects.filter(product_name=product_name,admin=admin).exists():
                return JsonResponse({
                "error": f"Duplicate product name '{product_name}' found in row {i}. Product name must be unique.",
                "products": product_responses
            }, status=400)
            product_name = product_name.lower()
            sku_number = row_data['sku_number']
            price = float(row_data['price'])
            quantity = int(row_data['quantity'])
            description = row_data['description']
            discount = int(row_data.get('discount') or 0)
            gst = float(row_data.get('gst') or 0)
            hsn_code = row_data.get('hsn_code', '')
            availability = "Out of Stock" if quantity == 0 else "Very Few Products Left" if quantity <= 10 else "In Stock"
            current_time = datetime.utcnow() + timedelta(hours=5, minutes=30)
            image_names = str(row_data.get('image_paths') or '').split(',')
            s3_image_keys = []
            for img_name in image_names:
                img_name = img_name.strip()
                if not img_name:
                    continue
                ext = img_name.split('.')[-1].lower()
                img_file = uploaded_images.get(img_name)
                if not img_file:
                    print(f"Uploaded images keys: {list(uploaded_images.keys())}")
                    return JsonResponse({
                        "error": f"Missing image file: '{img_name}' for product '{product_name}' in Excel row {i}. "
                                "Please ensure the image is selected in the upload form and the file name matches exactly.",
                                "products": product_responses,
                    }, status=400)
                s3_key = f"static/images/products/{product_name.replace(' ', '')}/{sku_number}{img_name}"
                try:
                    s3.upload_fileobj(
                        img_file,
                        settings.AWS_STORAGE_BUCKET_NAME,
                        s3_key,
                        ExtraArgs={'ContentType': f'image/{ext}'}
                    )
                except ClientError as e:
                    return JsonResponse({"error": f"Failed to upload image {img_name} on row {i}: {str(e)}"}, status=500)
                s3_image_keys.append(s3_key)
            material_name = str(row_data.get('material_paths') or '').strip()
            s3_material_key = None
            if material_name:
                ext = material_name.split('.')[-1].lower()
                material_file = uploaded_materials.get(material_name)
                if material_name and not material_file:
                    return JsonResponse({
                        "error": f"Missing material file: '{material_name}' for product '{product_name}' in Excel row {i}. "
                                "Please make sure the material file is included in the upload and named correctly.",
                                "products": product_responses,
                    }, status=400)
                s3_material_key = f"static/materials/{product_name.replace(' ', '_')}.{ext}"
                try:
                    s3.upload_fileobj(
                        material_file,
                        settings.AWS_STORAGE_BUCKET_NAME,
                        s3_material_key,
                        ExtraArgs={'ContentType': f'application/{ext}'}
                    )
                except ClientError as e:
                    return JsonResponse({"error": f"Failed to upload material on row {i}: {str(e)}","products": product_responses,}, status=500)
            specifications_str = str(row_data.get('specifications') or '').strip().rstrip(';')
            specs_list = [spec for spec in specifications_str.split(';') if spec]
            specifications_dict = {}
            for spec in specs_list:
                if ':' in spec:
                    key, value = spec.split(':', 1)
                    specifications_dict[key.strip()] = value.strip()
            number_of_specifications = len(specifications_dict)
            product = ProductsDetails.objects.create(
                product_name=product_name,
                sku_number=sku_number,
                hsn_code=hsn_code,
                price=price,
                quantity=quantity,
                discount=discount,
                gst=gst,
                description=description,
                admin=admin,
                category=category,
                sub_category=sub_category,
                availability=availability,
                created_at=current_time,
                product_status=1,
                cart_status=False,
                product_images=s3_image_keys,
                material_file=s3_material_key,
                specifications=specifications_dict,
                number_of_specifications=number_of_specifications
            )
            product_responses.append({
                "category_id":str(category_id),
                "su_category_id":str(sub_category_id),
                "product_id": str(product.id),
                "product_name": product.product_name,
                "sku_number": product.sku_number,
                "price": product.price,
                "quantity": product.quantity,
                "availability": product.availability,
                "specifications": product.specifications,
                "number_of_specifications": product.number_of_specifications
            })
        except Exception as e:
            return JsonResponse({"error": f"Error in row {i}: {str(e)}"}, status=500)
    return JsonResponse({
        "message": "All products uploaded successfully.",
        "products": product_responses,
        "admin_id":str(admin_id),
    }, status=200)
@csrf_exempt
def add_product_specifications(request):
    if request.method == 'POST':
        try:
            try:
                data = json.loads(request.body.decode('utf-8'))
            except json.JSONDecodeError:
                return JsonResponse({"error": "Invalid JSON format.", "status_code": 400}, status=400)
            admin_id = data.get('admin_id')
            category_id = data.get('category_id')
            sub_category_id = data.get('sub_category_id')
            product_id = data.get('product_id')
            new_specifications = data.get('specifications', [])
            if not all([admin_id, category_id, sub_category_id, product_id]):
                return JsonResponse({"error": "Missing required fields.", "status_code": 400}, status=400)
            try:
                admin = PavamanAdminDetails.objects.get(id=admin_id)
            except PavamanAdminDetails.DoesNotExist:
                return JsonResponse({"error": "Admin not found.", "status_code": 401}, status=401)
            try:
                category = CategoryDetails.objects.get(id=category_id, admin=admin)
            except CategoryDetails.DoesNotExist:
                return JsonResponse({"error": "Category not found for this admin.", "status_code": 404}, status=404)
            try:
                sub_category = SubCategoryDetails.objects.get(id=sub_category_id, category=category)
            except SubCategoryDetails.DoesNotExist:
                return JsonResponse({"error": "Subcategory not found for this category.", "status_code": 404}, status=404)
            try:
                product = ProductsDetails.objects.get(id=product_id, category=category, sub_category=sub_category)
            except ProductsDetails.DoesNotExist:
                return JsonResponse({"error": "Product not found.", "status_code": 404}, status=404)
            if not isinstance(new_specifications, list):
                return JsonResponse({"error": "Specifications must be a list of objects.", "status_code": 400}, status=400)
            existing_specifications = product.specifications or {}
            for spec in new_specifications:
                if "name" in spec and "value" in spec:
                    spec_name = spec["name"]
                    if spec_name in existing_specifications:
                        return JsonResponse({
                            "error": f"Specification '{spec_name}' already exists.",
                            "status_code": 400
                        }, status=400)
                    existing_specifications[spec_name] = spec["value"]
                else:
                    return JsonResponse({"error": "Each specification must contain 'name' and 'value'.", "status_code": 400}, status=400)
            product.specifications = existing_specifications
            product.number_of_specifications = len(existing_specifications)
            product.save()
            return JsonResponse({
                "message": "New specifications added successfully.",
                "product_id": str(product.id),
                "number_of_specifications": product.number_of_specifications,
                "specifications": product.specifications,
                "status_code": 200
            }, status=200)
        except Exception as e:
            return JsonResponse({"error": f"Unexpected error: {str(e)}", "status_code": 500}, status=500)
    return JsonResponse({"error": "Invalid request method. Only POST is allowed.", "status_code": 405}, status=405)
@csrf_exempt
def edit_product_specifications(request):
    if request.method == 'POST':
        try:
            try:
                data = json.loads(request.body.decode('utf-8'))
            except json.JSONDecodeError:
                return JsonResponse({"error": "Invalid JSON format.", "status_code": 400}, status=400)
            admin_id = data.get('admin_id')
            category_id = data.get('category_id')
            sub_category_id = data.get('sub_category_id')
            product_id = data.get('product_id')
            new_specifications = data.get('specifications', [])
            if not all([admin_id, category_id, sub_category_id, product_id]):
                return JsonResponse({"error": "Missing required fields.", "status_code": 400}, status=400)
            try:
                admin = PavamanAdminDetails.objects.get(id=admin_id)
            except PavamanAdminDetails.DoesNotExist:
                return JsonResponse({"error": "Admin not found.", "status_code": 401}, status=401)
            try:
                category = CategoryDetails.objects.get(id=category_id, admin=admin)
            except CategoryDetails.DoesNotExist:
                return JsonResponse({"error": "Category not found for this admin.", "status_code": 404}, status=404)
            try:
                sub_category = SubCategoryDetails.objects.get(id=sub_category_id, category=category)
            except SubCategoryDetails.DoesNotExist:
                return JsonResponse({"error": "Subcategory not found for this category.", "status_code": 404}, status=404)
            try:
                product = ProductsDetails.objects.get(id=product_id, category=category, sub_category=sub_category)
            except ProductsDetails.DoesNotExist:
                return JsonResponse({"error": "Product not found.", "status_code": 404}, status=404)
            if not isinstance(new_specifications, list):
                return JsonResponse({"error": "Specifications must be a list of objects.", "status_code": 400}, status=400)
            existing_specifications = product.specifications or {}
            for spec in new_specifications:
                if "name" in spec and "value" in spec:
                    existing_specifications[spec["name"]] = spec["value"]
                else:
                    return JsonResponse({"error": "Each specification must contain 'name' and 'value'.", "status_code": 400}, status=400)
            product.specifications = existing_specifications
            product.number_of_specifications = len(existing_specifications)
            product.save()
            return JsonResponse({
                "message": "Specifications updated successfully.",
                "product_id": str(product.id),
                "number_of_specifications": product.number_of_specifications,
                "specifications": product.specifications,
                "status_code": 200
            }, status=200)
        except Exception as e:
            return JsonResponse({"error": f"Unexpected error: {str(e)}", "status_code": 500}, status=500)
    return JsonResponse({"error": "Invalid request method. Only POST is allowed.", "status_code": 405}, status=405)
@csrf_exempt
def view_products(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body.decode('utf-8'))
            admin_id = data.get('admin_id')
            category_id = data.get('category_id')
            sub_category_id = data.get('sub_category_id')
            if not admin_id or not category_id or not sub_category_id:
                return JsonResponse({
                    "error": "admin_id, category_id, and sub_category_id are required.",
                    "status_code": 400
                }, status=400)
            try:
                admin = PavamanAdminDetails.objects.get(id=admin_id)
                category = CategoryDetails.objects.get(id=category_id, admin=admin)
                sub_category = SubCategoryDetails.objects.get(id=sub_category_id, category=category)
            except PavamanAdminDetails.DoesNotExist:
                return JsonResponse({"error": "Admin not found.", "status_code": 404}, status=404)
            except CategoryDetails.DoesNotExist:
                return JsonResponse({"error": "Category not found.", "status_code": 404}, status=404)
            except SubCategoryDetails.DoesNotExist:
                return JsonResponse({"error": "Subcategory not found.", "status_code": 404}, status=404)
            products = ProductsDetails.objects.filter(
                admin=admin, category=category, sub_category=sub_category
            ).values(
                'id', 'product_name', 'sku_number','hsn_code', 'price', 'availability', 'quantity', 'cart_status','product_images','discount','gst','description'
            )
            product_list = []
            for product in products:
                image_url = None
                if product['product_images']:
                    image_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{product['product_images'][0]}"
                price = round(float(product['price']), 2)
                discount = round(float(product.get('discount') or 0), 2)
                gst = round(float(product.get('gst') or 0), 2)
                discount_amount = round(price * (discount / 100), 2)
                final_price = round(price - discount_amount, 2)
                product_list.append({
                    "product_id": str(product['id']),
                    "product_name": product['product_name'],
                    "sku_number": product['sku_number'],
                    "hsn_code": product['hsn_code'],          
                    "price": price,
                    "availability": product['availability'],
                    "quantity": product['quantity'],
                    "cart_status":product['cart_status'],
                    "product_images": image_url,
                    "gst": int(gst),
                    "final_price": f"{final_price:.2f}",
                    "product_discount": int(discount),
                    "product_description":product['description']
                })
            return JsonResponse({
                "message": "Products retrieved successfully.",
                "status_code": 200,
                "category_id": str(category.id),
                "category_name": category.category_name,
                "sub_category_id": str(sub_category.id),
                "sub_category_name": sub_category.sub_category_name,
                "products": product_list
            }, status=200)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON format.", "status_code": 400}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"An unexpected error occurred: {str(e)}", "status_code": 500}, status=500)
    return JsonResponse({"error": "Invalid HTTP method. Only POST is allowed.", "status_code": 405}, status=405)
@csrf_exempt
def view_product_details(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body.decode('utf-8'))
            admin_id = data.get('admin_id')
            category_id = data.get('category_id')
            sub_category_id = data.get('sub_category_id')
            product_id = data.get('product_id')
            if not all([admin_id, category_id, sub_category_id, product_id]):
                return JsonResponse({
                    "error": "admin_id, category_id, sub_category_id, and product_id are required.",
                    "status_code": 400
                }, status=400)
            try:
                admin = PavamanAdminDetails.objects.get(id=admin_id)
                category = CategoryDetails.objects.get(id=category_id, admin=admin)
                sub_category = SubCategoryDetails.objects.get(id=sub_category_id, category=category)
                product = ProductsDetails.objects.get(id=product_id, admin=admin, category=category, sub_category=sub_category)
            except PavamanAdminDetails.DoesNotExist:
                return JsonResponse({"error": "Admin not found.", "status_code": 404}, status=404)
            except CategoryDetails.DoesNotExist:
                return JsonResponse({"error": "Category not found.", "status_code": 404}, status=404)
            except SubCategoryDetails.DoesNotExist:
                return JsonResponse({"error": "Subcategory not found.", "status_code": 404}, status=404)
            except ProductsDetails.DoesNotExist:
                return JsonResponse({"error": "Product not found.", "status_code": 404}, status=404)
            price = round(float(product.price), 2)
            discount = round(float(product.discount or 0))
            gst = round(float(product.gst or 0))
            discount_amount = round(price * (discount / 100), 2)
            final_price = round(price - discount_amount, 2)
            image_urls = []
            if product.product_images:
                for image in product.product_images:
                    image_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{image}"
                    image_urls.append(image_url)
            product_data = {
                "product_id": str(product.id),
                "product_name": product.product_name,
                "sku_number": product.sku_number,
                "hsn_code": product.hsn_code,
                "price": f"{price:.2f}",
                "discount": f"{discount}%",
                "discount_amount": f"{discount_amount:.2f}",
                "final_price": f"{final_price:.2f}",
                "gst": f"{gst}%",
                "availability": product.availability,
                "quantity": product.quantity,
                "description": product.description,
                "product_images": image_urls,
                "material_file": f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{product.material_file}",
                "number_of_specifications": product.number_of_specifications,
                "specifications": product.specifications,
            }
            return JsonResponse({
                "message": "Product details retrieved successfully.",
                "status_code": 200,
                "category_id": str(category.id),
                "category_name": category.category_name,
                "sub_category_id": str(sub_category.id),
                "sub_category_name": sub_category.sub_category_name,
                "product_details": product_data
            }, status=200)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON format.", "status_code": 400}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"An unexpected error occurred: {str(e)}", "status_code": 500}, status=500)

    return JsonResponse({"error": "Invalid HTTP method. Only POST is allowed.", "status_code": 405}, status=405)
@csrf_exempt
def edit_product(request):
    if request.method == 'POST':
        try:
            if request.content_type == "application/json":
                try:
                    data = json.loads(request.body.decode('utf-8'))
                except json.JSONDecodeError:
                    return JsonResponse({"error": "Invalid JSON format.", "status_code": 400}, status=400)
            else:
                data = request.POST.dict()

            admin_id = data.get('admin_id')
            category_id = data.get('category_id')
            sub_category_id = data.get('sub_category_id')
            product_id = data.get('product_id')
            product_name = data.get('product_name').lower()
            sku_number = data.get('sku_number')
            hsn_code = data.get('hsn_code')
            price = data.get('price')
            quantity = data.get('quantity')
            discount = data.get('discount', 0.0)
            description = data.get('description')
            gst = float(data.get('gst', 0.0))
            if not all([admin_id, category_id, sub_category_id, product_id, product_name, sku_number, hsn_code, price, quantity, description]):
                return JsonResponse({"error": "Missing required fields.", "status_code": 400}, status=400)
            try:
                price = float(price)
                quantity = int(quantity)
                discount = int(discount)
            except ValueError:
                return JsonResponse({"error": "Invalid format for price, quantity, or discount.", "status_code": 400}, status=400)
            if discount > price:
                return JsonResponse({"error": "Discount cannot be greater than the price.", "status_code": 400}, status=400)
            availability = "In Stock" if quantity > 10 else "Very Few Products Left" if quantity > 0 else "Out of Stock"
            try:
                admin = PavamanAdminDetails.objects.get(id=admin_id)
            except PavamanAdminDetails.DoesNotExist:
                return JsonResponse({"error": "Admin not found.", "status_code": 401}, status=401)
            try:
                category = CategoryDetails.objects.get(id=category_id, admin=admin)
            except CategoryDetails.DoesNotExist:
                return JsonResponse({"error": "Category not found.", "status_code": 404}, status=404)
            try:
                sub_category = SubCategoryDetails.objects.get(id=sub_category_id, category=category)
            except SubCategoryDetails.DoesNotExist:
                return JsonResponse({"error": "Subcategory not found.", "status_code": 404}, status=404)
            try:
                product = ProductsDetails.objects.get(id=product_id, category=category, sub_category=sub_category)
            except ProductsDetails.DoesNotExist:
                return JsonResponse({"error": "Product not found.", "status_code": 404}, status=404)
            if ProductsDetails.objects.exclude(id=product_id).filter(sku_number=sku_number).exists():
                return JsonResponse({"error": "SKU number already exists.", "status_code": 400}, status=400)
            if ProductsDetails.objects.exclude(id=product_id).filter(product_name=product_name).exists():
                return JsonResponse({"error": "Product name already exists.", "status_code": 400}, status=400)
            old_product_name = product.product_name
            old_sku_number = product.sku_number
            product.product_name = product_name
            product.sku_number = sku_number
            product.hsn_code = hsn_code
            product.price = price
            product.quantity = quantity
            product.discount = discount
            product.gst = gst
            product.description = description
            product.availability = availability
            product.cart_status = False
            s3 = boto3.client(
                's3',
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                region_name=settings.AWS_S3_REGION_NAME
            )
            product_images = product.product_images or []

            old_folder = f"static/images/products/{old_product_name.replace(' ', '_').replace('/', '_')}"
            new_folder = f"static/images/products/{product_name.replace(' ', '_').replace('/', '_')}"

            if old_product_name != product_name or old_sku_number != sku_number:
                temp_images = []
                try:
                    response = s3.list_objects_v2(Bucket=settings.AWS_STORAGE_BUCKET_NAME, Prefix=old_folder)
                    if 'Contents' in response:
                        for obj in response['Contents']:
                            old_key = obj['Key']
                            filename = old_key.split('/')[-1]
                            if old_sku_number in filename:
                                new_filename = filename.replace(old_sku_number, sku_number)
                            else:
                                new_filename = f"{sku_number}_{filename}"
                            new_key = f"{new_folder}/{new_filename}"
                            s3.copy_object(
                                Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                                CopySource={'Bucket': settings.AWS_STORAGE_BUCKET_NAME, 'Key': old_key},
                                Key=new_key,
                                ContentType=obj.get('ContentType', 'image/jpeg')
                            )
                            temp_images.append(new_key)
                        s3.delete_objects(Bucket=settings.AWS_STORAGE_BUCKET_NAME, Delete={
                            'Objects': [{'Key': obj['Key']} for obj in response['Contents']]
                        })
                        product_images = temp_images
                except Exception as move_err:
                    print("Warning: Failed to rename/move images:", str(move_err))
            if 'product_images' in request.FILES:
                product_images = []
                try:
                    response = s3.list_objects_v2(Bucket=settings.AWS_STORAGE_BUCKET_NAME, Prefix=new_folder)
                    if 'Contents' in response:
                        s3.delete_objects(Bucket=settings.AWS_STORAGE_BUCKET_NAME, Delete={
                            'Objects': [{'Key': obj['Key']} for obj in response['Contents']]
                        })
                except Exception as del_err:
                    print("Warning: Failed to delete existing images:", str(del_err))
  
                image_files = request.FILES.getlist('product_images')
                for image in image_files:
                    allowed_extensions = ['png', 'jpg', 'jpeg']
                    file_extension = image.name.split('.')[-1].lower()
                    if file_extension not in allowed_extensions:
                        return JsonResponse({"error": f"Invalid file type. Allowed types: {', '.join(allowed_extensions)}", "status_code": 400}, status=400)
                    safe_image_name = f"{sku_number}_{image.name.replace(' ', '_').replace('/', '_')}"
                    s3_file_key = f"{new_folder}/{safe_image_name}"
                    try:
                        s3.upload_fileobj(
                            image,
                            settings.AWS_STORAGE_BUCKET_NAME,
                            s3_file_key,
                            ExtraArgs={'ContentType': image.content_type}
                        )
                        product_images.append(s3_file_key)
                    except Exception as e:
                        return JsonResponse({"error": f"Failed to upload image to S3: {str(e)}", "status_code": 500}, status=500)
            product.product_images = product_images
            if old_product_name != product_name and product.material_file:
                old_material_key = product.material_file
                file_extension = old_material_key.split('.')[-1]
                new_material_key = f"static/materials/{product_name}.{file_extension}"
                try:
                    s3.copy_object(
                        Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                        CopySource={'Bucket': settings.AWS_STORAGE_BUCKET_NAME, 'Key': old_material_key},
                        Key=new_material_key,
                        ContentType='application/pdf' if file_extension == 'pdf' else 'application/msword'
                    )
                    s3.delete_object(Bucket=settings.AWS_STORAGE_BUCKET_NAME, Key=old_material_key)
                    product.material_file = new_material_key
                except Exception as rename_err:
                    print("Warning: Failed to rename material file:", str(rename_err))
            if 'material_file' in request.FILES:
                material_file = request.FILES['material_file']
                allowed_extensions = ['pdf', 'doc']
                file_extension = material_file.name.split('.')[-1].lower()
                if file_extension not in allowed_extensions:
                    return JsonResponse({"error": f"Invalid material file type. Allowed types: {', '.join(allowed_extensions)}", "status_code": 400}, status=400)
                safe_material_name = f"{product_name}.{file_extension}"
                s3_material_key = f"static/materials/{safe_material_name}"
                try:
                    s3.upload_fileobj(
                        material_file,
                        settings.AWS_STORAGE_BUCKET_NAME,
                        s3_material_key,
                        ExtraArgs={'ContentType': material_file.content_type}
                    )
                    product.material_file = s3_material_key
                except Exception as e:
                    return JsonResponse({"error": f"Failed to upload material file to S3: {str(e)}", "status_code": 500}, status=500)
            final_price = price - (price * discount / 100)
            product.save()
            return JsonResponse({
                "message": "Product updated successfully.",
                "category_id": str(product.category.id),
                "category_name": product.category.category_name,
                "subcategory_id": str(product.sub_category.id),
                "sub_category_name": product.sub_category.sub_category_name,
                "product_id": str(product.id),
                "availability": availability,
                "cart_status": product.cart_status,
                "price": round(price, 2),
                "discount": round(discount, 2),
                "gst": round(gst, 2),
                "hsn_code": hsn_code,
                "sku_number": sku_number,
                "final_price": round(final_price, 2),
                "status_code": 200
            }, status=200)

        except Exception as e:
            return JsonResponse({"error": f"Unexpected error: {str(e)}", "status_code": 500}, status=500)
    else:
        return JsonResponse({"error": "Invalid request method. Only POST is allowed.", "status_code": 405}, status=405)
@csrf_exempt
def delete_product(request):
    if request.method == 'POST':
        try:
            if request.content_type == "application/json":
                try:
                    data = json.loads(request.body.decode('utf-8'))
                except json.JSONDecodeError:
                    return JsonResponse({"error": "Invalid JSON format.", "status_code": 400}, status=400)
            else:
                data = request.POST.dict()
            admin_id = data.get('admin_id')
            category_id = data.get('category_id')
            sub_category_id = data.get('sub_category_id')
            product_id = data.get('product_id')
            if not all([admin_id, category_id, sub_category_id, product_id]):
                return JsonResponse({"error": "Missing required fields.", "status_code": 400}, status=400)
            try:
                admin = PavamanAdminDetails.objects.get(id=admin_id)
            except PavamanAdminDetails.DoesNotExist:
                return JsonResponse({"error": "Admin not found.", "status_code": 401}, status=401)
            try:
                category = CategoryDetails.objects.get(id=category_id, admin=admin)
            except CategoryDetails.DoesNotExist:
                return JsonResponse({"error": "Category not found.", "status_code": 404}, status=404)
            try:
                sub_category = SubCategoryDetails.objects.get(id=sub_category_id, category=category)
            except SubCategoryDetails.DoesNotExist:
                return JsonResponse({"error": "Sub-category not found.", "status_code": 404}, status=404)
            try:
                product = ProductsDetails.objects.get(id=product_id, category=category, sub_category=sub_category)
            except ProductsDetails.DoesNotExist:
                return JsonResponse({"error": "Product not found.", "status_code": 404}, status=404)
            s3 = boto3.client(
                's3',
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                region_name=settings.AWS_S3_REGION_NAME
            )
            product_folder_prefix = f"static/images/products/{product.product_name.replace(' ', '_')}/"
            try:
                response = s3.list_objects_v2(
                    Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                    Prefix=product_folder_prefix
                )
                if 'Contents' in response:
                    for item in response['Contents']:
                        s3.delete_object(Bucket=settings.AWS_STORAGE_BUCKET_NAME, Key=item['Key'])
            except ClientError as e:
                print(f"Error deleting product folder from S3: {e}")
            if product.material_file:
                try:
                    s3.delete_object(
                        Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                        Key=product.material_file
                    )
                except ClientError as e:
                    print(f"Error deleting material file from S3: {e}")
            product.delete()
            return JsonResponse({
                "message": "Product and associated files deleted successfully.",
                "product_id": product_id,
                "status_code": 200
            }, status=200)
        except Exception as e:
            return JsonResponse({"error": f"Unexpected error: {str(e)}", "status_code": 500}, status=500)
    return JsonResponse({"error": "Invalid request method. Only POST is allowed.", "status_code": 405}, status=405)
@csrf_exempt
def search_categories(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            admin_id = data.get('admin_id')
            search_query = data.get('category_name', '').strip()
            if not admin_id:
                return JsonResponse({"error": "Admin Id is required.", "status_code": 400}, status=400)
            if not search_query:
                return JsonResponse({"error": "Atleast one character is required.", "status_code": 400}, status=400)
            admin_data = PavamanAdminDetails.objects.filter(id=admin_id).first()
            if not admin_data:
                return JsonResponse({"error": "Admin not found or session expired.", "status_code": 401}, status=401)
            categories = CategoryDetails.objects.filter(
                admin_id=admin_id,
                category_status=1,
                category_name__icontains=search_query
            )
            if not categories.exists():
                return JsonResponse({"message": "No category details found", "status_code": 200}, status=200)
            category_list = []
            for category in categories:
                image_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{category.category_image}"
                category_list.append({
                    "category_id": str(category.id),
                    "category_name": category.category_name,
                    "category_image_url": image_url
                })
            return JsonResponse(
                {"message": "Categories retrieved successfully.", "categories": category_list, "status_code": 200},
                status=200
            )
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON data.", "status_code": 400}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"An unexpected error occurred: {str(e)}", "status_code": 500}, status=500)
    return JsonResponse({"error": "Invalid HTTP method. Only POST is allowed.", "status_code": 405}, status=405)
@csrf_exempt
def search_subcategories(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            admin_id = data.get('admin_id')
            category_id = data.get('category_id')
            sub_category_name = data.get('sub_category_name', '').strip()
            if not admin_id:
                return JsonResponse({"error": "Admin Id is required.", "status_code": 400}, status=400)
            if not category_id:
                return JsonResponse({"error": "Category Id is required.", "status_code": 400}, status=400)
            if sub_category_name == "": 
                return JsonResponse({"error": "Atleast one character is required.", "status_code": 400}, status=400)
            admin_data = PavamanAdminDetails.objects.filter(id=admin_id).first()
            if not admin_data:
                return JsonResponse({"error": "Admin not found or session expired.", "status_code": 401}, status=401)
            subcategories = SubCategoryDetails.objects.filter(
                admin_id=admin_id,
                category_id=category_id,
                sub_category_status=1,
                sub_category_name__icontains=sub_category_name 
            )
            if not subcategories.exists():
                return JsonResponse({"message": "No subcategory details found", "status_code": 200}, status=200)
            subcategory_list = []
            for subcategory in subcategories:
                image_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{subcategory.sub_category_image}"
                subcategory_list.append({
                    "sub_category_id": str(subcategory.id),
                    "sub_category_name": subcategory.sub_category_name,
                    "sub_category_image_url": image_url
                })
            return JsonResponse(
                {"message": "Subcategories retrieved successfully.", "subcategories": subcategory_list, "status_code": 200},
                status=200
            )
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON data.", "status_code": 400}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"An unexpected error occurred: {str(e)}", "status_code": 500}, status=500)
    return JsonResponse({"error": "Invalid HTTP method. Only POST is allowed.", "status_code": 405}, status=405)
@csrf_exempt
def search_products(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            admin_id = data.get('admin_id')
            category_id = data.get('category_id')
            sub_category_id = data.get('sub_category_id')
            product_name = data.get('product_name', '').strip() 
            if not admin_id:
                return JsonResponse({"error": "Admin ID are required.", "status_code": 400}, status=400)
            if not category_id:
                return JsonResponse({"error": "Category ID are required.", "status_code": 400}, status=400)
            if not sub_category_id:
                return JsonResponse({"error": "Sub Category ID are required.", "status_code": 400}, status=400)
            if product_name == "":
                return JsonResponse({"error": "Atleast one character is required.", "status_code": 400}, status=400)
            admin_data = PavamanAdminDetails.objects.filter(id=admin_id).first()
            if not admin_data:
                return JsonResponse({"error": "Admin not found or session expired.", "status_code": 401}, status=401)
            products = ProductsDetails.objects.filter(
                admin_id=admin_id,
                category_id=category_id,
                sub_category_id=sub_category_id,
                product_status=1
            ).filter(
                Q(product_name__icontains=product_name) |
                Q(sku_number__icontains=product_name)
            )
            if not products.exists():
                return JsonResponse({"message": "No product details found", "status_code": 200}, status=200)
            product_list = []
            for product in products:
                product_images = product.product_images
                if isinstance(product_images, list) and product_images:
                    product_image_key = product_images[0]
                elif isinstance(product_images, str):
                    product_image_key = product_images
                else:
                    product_image_key = ""

                product_image_url = (
                    f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{product_image_key}"
                    if product_image_key else ""
                )
                price = round(float(product.price), 2)
                discount = round(float(product.discount or 0), 2)
                gst = round(float(product.gst or 0), 2)
                discount_amount = round(price * (discount / 100), 2)
                final_price = round(price - discount_amount, 2)
                product_list.append({
                    "product_id": str(product.id),
                    "product_name": product.product_name,
                    "category_id": str(product.category_id),
                    "sub_category_id": str(product.sub_category_id),
                    "product_images": product_image_url,
                    "sku_number": product.sku_number,
                    "hsn":product.hsn_code,
                    "price":product.price,
                    "gst": f"{int(gst)}%",
                    "final_price": f"{final_price:.2f}",
                    "product_discount": f"{int(discount)}%"
                })
            return JsonResponse(
                {"message": "Products retrieved successfully.", "products": product_list, "status_code": 200},
                status=200
            )
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON data.", "status_code": 400}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"An unexpected error occurred: {str(e)}", "status_code": 500}, status=500)
    return JsonResponse({"error": "Invalid HTTP method. Only POST is allowed.", "status_code": 405}, status=405)
@csrf_exempt
def download_discount_products_excel(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            admin_id = data.get('admin_id')
            if not admin_id:
                return JsonResponse({"error": "Admin ID is required.", "status_code": 400}, status=400)
            products = ProductsDetails.objects.filter(admin_id=admin_id).order_by('created_at')
            if not products.exists():
                return JsonResponse({
                    "message": "No products with discount found.",
                    "status_code": 200,
                    "admin_id": str(admin_id)
                }, status=200)
            wb = Workbook()
            ws = wb.active
            ws.title = "Products Details"
            headers = [
                "Product ID", "Product Name", "SKU","HSN", "Price", "Discount (%)", "GST (%)","Final Price",
                "Quantity", "Material File", "Description", "Specifications Count", "Specifications",
                "Availability", "Category", "Subcategory", "Created At"
            ]
            ws.append(headers)
            for product in products:
                price = float(product.price)
                discount= float(product.discount)
                gst = float(product.gst) if hasattr(product, 'gst') and product.gst else 0
                final_price = round(price - (price * discount / 100), 2)
                material_file_url = (
                    f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{product.material_file}"
                    if product.material_file else ""
                )
                ws.append([
                    str(product.id),
                    product.product_name,
                    product.sku_number,
                    product.hsn_code,
                    round(price, 2),
                    f"{round(discount)}%",
                    f"{round(gst)}%",
                    round(final_price, 2),
                    product.quantity,
                    material_file_url,
                    product.description,
                    product.number_of_specifications,
                    json.dumps(product.specifications) if isinstance(product.specifications, dict) else product.specifications,
                    product.availability,
                    product.category.category_name if product.category else '',
                    product.sub_category.sub_category_name if product.sub_category else '',
                    product.created_at.strftime('%Y-%m-%d %H:%M:%S') if product.created_at else ''
                ])
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            response = HttpResponse(
                buffer,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=Products_Details.xlsx'
            return response
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON data.", "status_code": 400}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"An unexpected error occurred: {str(e)}", "status_code": 500}, status=500)
    return JsonResponse({"error": "Invalid HTTP method. Only POST is allowed.", "status_code": 405}, status=405)
@csrf_exempt
def apply_discount_by_subcategory_only(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body.decode('utf-8'))
            categories = data.get('categories')
            admin_id = data.get('admin_id')
            if not admin_id:
                return JsonResponse({"error": "admin_id is required.", "status_code": 400}, status=400)
            try:
                admin = PavamanAdminDetails.objects.get(id=admin_id, status=1)
            except PavamanAdminDetails.DoesNotExist:
                return JsonResponse({"error": "Admin not found or not active.", "status_code": 403}, status=403)
            if not categories or not isinstance(categories, list):
                return JsonResponse({"error": "categories must be a list.", "status_code": 400}, status=400)
            response_categories = []
            for cat_data in categories:
                category_id = cat_data.get('category_id')
                category_name = cat_data.get('category_name')
                sub_category_id = cat_data.get('sub_category_id')
                sub_category_name = cat_data.get('sub_category_name')
                discount_str = cat_data.get('discount')
                if not all([category_id, category_name, sub_category_id, sub_category_name, discount_str]):
                    return JsonResponse({
                        "error": "Each item must have category_id, category_name, sub_category_id, sub_category_name, and discount.",
                        "status_code": 400
                    }, status=400)
                try:
                    discount = float(discount_str)
                    if discount <= 0 or discount > 100:
                        return JsonResponse({"error": "Discount must be between 1 and 100.Do not use "%".", "status_code": 400}, status=400)
                except ValueError:
                    return JsonResponse({"error": "Discount must be a number.", "status_code": 400}, status=400)
                try:
                    category = CategoryDetails.objects.get(id=category_id, category_name=category_name)
                except CategoryDetails.DoesNotExist:
                    return JsonResponse({"error": f"Category '{category_name}' not found.", "status_code": 404}, status=404)
                try:
                    subcategory = SubCategoryDetails.objects.get(
                        id=sub_category_id,
                        sub_category_name=sub_category_name,
                        category_id=category_id
                    )
                except SubCategoryDetails.DoesNotExist:
                    return JsonResponse({
                        "error": f"Subcategory '{sub_category_name}' not found in category '{category_name}'.",
                        "status_code": 404
                    }, status=404)
                products = ProductsDetails.objects.filter(
                    category_id=category_id,
                    sub_category_id=sub_category_id,
                    product_status=1
                )
                updated_products = []
                for product in products:
                    price = float(product.price or 0)
                    gst = float(product.gst or 0)
                    discount_amount = (price * discount / 100) if price > 0 else 0
                    final_price = round(price - discount_amount, 2)
                    product.discount = discount
                    product.save(update_fields=['discount'])
                    updated_products.append({
                        "product_id": str(product.id),
                        "product_name": product.product_name,
                        "price": round(price, 2),
                        "discount": f"{round(discount, 2)}%",
                        "gst": f"{round(gst, 2)}%",
                        "final_price": round(final_price, 2)
                    })
                response_categories.append({
                    "category_id": category_id,
                    "category_name": category_name,
                    "sub_category_id": sub_category_id,
                    "sub_category_name": sub_category_name,
                    "discount": f"{int(discount)}%",
                    "admin_id": admin_id,
                    "updated_products": updated_products
                })
            return JsonResponse({
                "categories": response_categories,
                "admin_id": admin_id,
                "status_code": 200
            }, status=200)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON format.", "status_code": 400}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Server error: {str(e)}", "status_code": 500}, status=500)
    return JsonResponse({"error": "Invalid HTTP method. Only POST is allowed.", "status_code": 405}, status=405)
@csrf_exempt
def order_or_delivery_status(request):
    if request.method != "POST":
        return JsonResponse({"error": "Only POST allowed", "status_code": 405}, status=405)
    try:
        data = json.loads(request.body.decode("utf-8"))
        admin_id = data.get("admin_id")
        customer_id =data.get("customer_id")
        product_order_id = data.get("product_order_id")
        action = data.get("action")
        single_order_product_id = data.get("single_order_product_id")
        if not all([admin_id, product_order_id,customer_id, action]):
            return JsonResponse({"error": "Missing required fields", "status_code": 400}, status=400)
        payment = PaymentDetails.objects.filter(admin_id=admin_id, product_order_id=product_order_id,customer_id=customer_id).first()
        if not payment:
            return JsonResponse({"error": "Payment not found", "status_code": 404}, status=404)
        updated_orders = []
        if action == "Shipped":
            if single_order_product_id:
                if single_order_product_id not in payment.order_product_ids:
                    return JsonResponse({"error": "Invalid order product ID", "status_code": 404}, status=404)
                order = OrderProducts.objects.filter(id=single_order_product_id,customer_id=customer_id).first()
                if order:
                    order.shipping_status = "Shipped"
                    order.save()
                    updated_orders.append(order.id)
            else:
                for oid in payment.order_product_ids:
                    order = OrderProducts.objects.filter(id=oid,customer_id=customer_id).first()
                    if order:
                        order.shipping_status = "Shipped"
                        order.save()
                        updated_orders.append(order.id)
        elif action == "Delivered":
            if single_order_product_id:
                if single_order_product_id not in payment.order_product_ids:
                    return JsonResponse({"error": "Invalid order product ID", "status_code": 404}, status=404)
                order = OrderProducts.objects.filter(id=single_order_product_id, customer_id=customer_id).first()
                if order:
                    if order.shipping_status != "Shipped":
                        return JsonResponse({"error": "Cannot mark as Delivered before Shipped", "status_code": 400}, status=400)
                    order.delivery_status = "Delivered"
                    order.save()
                    updated_orders.append(order.id)
            else:
                for oid in payment.order_product_ids:
                    order = OrderProducts.objects.filter(id=oid, customer_id=customer_id).first()
                    if order:
                        if order.shipping_status != "Shipped":
                            return JsonResponse({
                                "error": f"OrderProduct ID {oid} has not been shipped yet",
                                "status_code": 400
                            }, status=400)
                        order.delivery_status = "Delivered"
                        order.save()
                        updated_orders.append(order.id)
        else:
            return JsonResponse({"error": "Invalid action type", "status_code": 400}, status=400)
        return JsonResponse({
            "message": f"{action.capitalize()} status updated successfully.",
            "updated_orders": updated_orders,
            "admin_id":str(admin_id),
            "status_code": 200
        })

    except Exception as e:
        return JsonResponse({"error": str(e), "status_code": 500}, status=500)
@csrf_exempt
def retrieve_feedback(request):
    if request.method != "POST":
        return JsonResponse({
            "error": "Invalid HTTP method. Only POST allowed.",
            "status_code": 405
        }, status=405)
    try:
        data = json.loads(request.body.decode("utf-8"))
        admin_id = data.get('admin_id')
        action = data.get('action')
        if not admin_id:
            return JsonResponse({"error": "admin_id is required.", "status_code": 400}, status=400)
        if not action:
            return JsonResponse({"error": "action is required.", "status_code": 400}, status=400)
        if action == "customer_rating":
            feedbacks = FeedbackRating.objects.filter(admin_id=admin_id)
            if not feedbacks.exists():
                return JsonResponse({"error": "No feedback found for this admin.", "status_code": 404}, status=404)
            feedback_data = []
            for feedback in feedbacks:
                try:
                    customer = CustomerRegisterDetails.objects.get(id=feedback.customer_id)
                    product = ProductsDetails.objects.get(id=feedback.product_id)
                    image_url = None
                    if product.product_images and isinstance(product.product_images, list):
                        first_image = product.product_images[0]
                        if first_image:
                            image_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{first_image}"
                    feedback_data.append({
                        "customer_id": customer.id,
                        "customer_name": f"{customer.first_name} {customer.last_name}",
                        "customer_email": customer.email,
                        "product_image": image_url,
                        "product_name": product.product_name,
                        "product_id": product.id,
                        "rating": feedback.rating,
                        "feedback": feedback.feedback,
                        "order_id": feedback.order_id,
                        "created_at": feedback.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                    })
                except CustomerRegisterDetails.DoesNotExist:
                    continue
            return JsonResponse({
                "feedback": feedback_data,
                "status_code": 200,
                "admin_id": str(admin_id)
            }, status=200)
        elif action == "avgrating":
            avg_ratings = FeedbackRating.objects.filter(admin_id=admin_id) \
                            .values('product_id') \
                            .annotate(average_rating=Avg('rating'))
            rating_data = []
            for item in avg_ratings:
                try:
                    product = ProductsDetails.objects.get(id=item['product_id'])
                    image_url = None
                    if product.product_images and isinstance(product.product_images, list):
                        first_image = product.product_images[0]
                        if first_image:
                            image_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{first_image}"
                    rating_data.append({
                        "product_id": product.id,
                        "category_name":product.category.category_name,
                        "subcategory_name":product.sub_category.sub_category_name,
                        "product_name": product.product_name,
                        "product_image": image_url,
                        "average_rating": round(item['average_rating'], 2),
                        "SKU":product.sku_number,
                        'HSN':product.hsn_code
                    })
                except ProductsDetails.DoesNotExist:
                    continue

            return JsonResponse({
                "average_ratings": rating_data,
                "status_code": 200,
                "admin_id": str(admin_id)
            }, status=200)

        else:
            return JsonResponse({"error": "Invalid action provided.", "status_code": 400}, status=400)
    except json.JSONDecodeError:
        return JsonResponse({
            "error": "Invalid JSON format.",
            "status_code": 400
        }, status=400)
    except Exception as e:
        return JsonResponse({
            "error": f"Server error: {str(e)}",
            "status_code": 500
        }, status=500)
@csrf_exempt
def report_inventory_summary(request):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid HTTP method. Only POST is allowed.", "status_code": 405}, status=405)
    try:
        data = json.loads(request.body.decode("utf-8"))
        admin_id = data.get('admin_id')
        if not admin_id:
            return JsonResponse({"error": "admin_id is required.", "status_code": 400}, status=400)
        total_products = ProductsDetails.objects.filter(admin_id=admin_id).count()
        total_customers = CustomerRegisterDetails.objects.filter(admin_id=admin_id).count()
        low_stock_products = ProductsDetails.objects.filter(
            admin_id=admin_id,
            quantity__lt=10
        ).values('product_name', 'sku_number', 'quantity')
        return JsonResponse({
            "total_products": total_products,
            "total_customers": total_customers,
            "low_stock_products": list(low_stock_products),
            "status_code": 200,
            "admin_id": admin_id
        }, status=200)
    except Exception as e:
        return JsonResponse({"error": str(e), "status_code": 500}, status=500)
@csrf_exempt
def top_buyers_report(request):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid HTTP method. Only POST is allowed.", "status_code": 405}, status=405)
    try:
        data = json.loads(request.body.decode("utf-8"))
        admin_id = data.get('admin_id')
        if not admin_id:
            return JsonResponse({"error": "admin_id is required.", "status_code": 400}, status=400)
        payments = PaymentDetails.objects.filter(admin_id=admin_id)
        all_order_product_ids = []
        for p in payments:
            if isinstance(p.order_product_ids, list):
                all_order_product_ids.extend(p.order_product_ids)
        buyers_data = (
            OrderProducts.objects
            .filter(admin_id=admin_id, id__in=all_order_product_ids)
            .values('customer_id')
            .annotate(
                product_count=Count('product_id'),      
                total_quantity=Sum('quantity')          
            )
            .order_by('-total_quantity')  
        )
        result = []
        for buyer in buyers_data:
            try:
                customer = CustomerRegisterDetails.objects.get(id=buyer['customer_id'])
                result.append({
                    "customer_id": customer.id,
                    "name": f"{customer.first_name} {customer.last_name}",
                    "email": customer.email,
                    "mobile_no": customer.mobile_no,
                    "product_count": buyer['product_count'],
                    "total_quantity": buyer['total_quantity']
                })
            except CustomerRegisterDetails.DoesNotExist:
                continue
        return JsonResponse({
            "buyers": result,
            "status_code": 200,
            "admin_id": admin_id
        }, status=200)
    except Exception as e:
        return JsonResponse({"error": str(e), "status_code": 500}, status=500)
@csrf_exempt
def monthly_product_orders(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body.decode('utf-8'))
            admin_id = data.get("admin_id")
            if not admin_id:
                return JsonResponse({
                    "status_code": 400,
                    "message": "admin_id is required."
                })
            now = timezone.now()
            start_of_month = datetime(now.year, now.month, 1, tzinfo=timezone.get_current_timezone())
            paid_order_ids = PaymentDetails.objects.filter(
                admin_id=admin_id,
                razorpay_payment_id__isnull=False
            ).values_list("order_product_ids", flat=True)
            order_ids = []
            for item in paid_order_ids:
                order_ids.extend(item)
            monthly_data = OrderProducts.objects.filter(
                admin_id=admin_id,
                id__in=order_ids,
                created_at__gte=start_of_month
            ).annotate(
                month=TruncMonth('created_at')
            ).values(
                'month'
            ).annotate(
                total_quantity=Sum('quantity')
            ).order_by('month')

            result = [
                {
                    "month": item["month"].strftime("%Y-%m"),
                    "total_quantity": item["total_quantity"]
                }
                for item in monthly_data
            ]
            return JsonResponse({
                "admin_id":admin_id,
                "status_code": 200,
                "message": "Monthly total products ordered (current month).",
                "data": result
            })
        except Exception as e:
            return JsonResponse({
                "status_code": 500,
                "message": "Error occurred.",
                "error": str(e)
            })
    else:
        return JsonResponse({
            "status_code": 405,
            "message": "Method Not Allowed. Use POST."
        }, status=405)
@csrf_exempt
def download_feedback_excel(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body.decode("utf-8"))
            admin_id = data.get("admin_id")

            if not admin_id:
                return JsonResponse({"error": "admin_id is required", "status_code": 400}, status=400)

            feedbacks = FeedbackRating.objects.filter(admin_id=admin_id).order_by("created_at")
            if not feedbacks.exists():
                return JsonResponse({"error": "No feedback found for this admin", "status_code": 404}, status=404)
            wb = Workbook()
            ws = wb.active
            ws.title = "Feedback Report"
            headers = [
                "Customer Name", "Customer Email", "Product Name", 
                "SKU","HSN", "Rating", "Feedback", "Order ID", 
                "Created At", "Product Image URL"
            ]
            ws.append(headers)
            for feedback in feedbacks:
                try:
                    customer = CustomerRegisterDetails.objects.get(id=feedback.customer_id)
                    product = ProductsDetails.objects.get(id=feedback.product_id)

                    image_url = None
                    if product.product_images and isinstance(product.product_images, list):
                        first_image = product.product_images[0]
                        if first_image:
                            image_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{first_image}"
                    ws.append([
                        f"{customer.first_name} {customer.last_name}",
                        customer.email,
                        product.product_name,
                        product.sku_number,
                        product.hsn_code,
                        feedback.rating,
                        feedback.feedback,
                        feedback.order_id,
                        feedback.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                        image_url or "",
                    ])
                except (CustomerRegisterDetails.DoesNotExist, ProductsDetails.DoesNotExist):
                    continue
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=feedback_report.xlsx'
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            response.write(buffer.read())
            return response
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON format", "status_code": 400}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Server error: {str(e)}", "status_code": 500}, status=500)
    return JsonResponse({"error": "Invalid HTTP method. Only POST allowed.", "status_code": 405}, status=405)
@csrf_exempt
def download_inventory_products_excel(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            admin_id = data.get('admin_id')
            if not admin_id:
                return JsonResponse({"error": "Admin ID is required.", "status_code": 400}, status=400)
            products = ProductsDetails.objects.filter(admin_id=admin_id)
            if not products.exists():
                return JsonResponse({
                    "message": "No products found.",
                    "status_code": 200,
                    "admin_id": str(admin_id)
                }, status=200)
            wb = Workbook()
            ws = wb.active
            ws.title = "Products Inventory Details"
            headers = [
                "Product Name", "SKU","HSN" ,"Price", "Discount (%)", "GST (%)", "Final Price",
                "Quantity", "Total Sold Quantity", "Specifications",
            ]
            ws.append(headers)
            for product in products:
                price = float(product.price)
                discount = float(product.discount)
                gst = float(product.gst) if hasattr(product, 'gst') and product.gst else 0
                final_price = round(price - discount, 2) if price > 0 else 0
                total_sold = OrderProducts.objects.filter(
                    product_id=product.id,
                    order_status='Paid'
                ).aggregate(total=Sum('quantity'))['total'] or 0
                ws.append([
                    product.product_name,
                    product.sku_number,
                    product.hsn_code,
                    round(price, 2),
                    f"{round(discount)}%",
                    f"{round(gst)}%",
                    round(final_price, 2),
                    product.quantity,
                    total_sold,
                    json.dumps(product.specifications) if isinstance(product.specifications, dict) else product.specifications,
                ])
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            response = HttpResponse(
                buffer,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=Products_Inventory_Details.xlsx'
            return response
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON data.", "status_code": 400}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"An unexpected error occurred: {str(e)}", "status_code": 500}, status=500)
    return JsonResponse({"error": "Invalid HTTP method. Only POST is allowed.", "status_code": 405}, status=405)
@csrf_exempt
def download_average_rating_excel(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body.decode("utf-8"))
            admin_id = data.get("admin_id")

            if not admin_id:
                return JsonResponse({"error": "admin_id is required", "status_code": 400}, status=400)
            avg_ratings = list(
            FeedbackRating.objects.filter(admin_id=admin_id)
            .values('product_id', 'created_at')
            .annotate(average_rating=Avg('rating'))
            )
            avg_ratings.sort(key=lambda x: x['created_at'])

            if not avg_ratings:
                return JsonResponse({"error": "No ratings found for this admin", "status_code": 404}, status=404)
            wb = Workbook()
            ws = wb.active
            ws.title = "Average Ratings"
            headers = ["Product Name","SKU","HSN", "Average Rating","Category Name","SubCategory Name"]
            ws.append(headers)
            for item in avg_ratings:
                try:
                    product = ProductsDetails.objects.get(id=item['product_id'])
                    ws.append([
                        product.product_name,
                        product.sku_number,
                        product.hsn_code,                     
                        round(item['average_rating'], 2),
                        product.category.category_name,
                        product.sub_category.sub_category_name,
                    ])
                except ProductsDetails.DoesNotExist:
                    continue
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=product_rating_report.xlsx'
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            response.write(buffer.read())
            return response
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON format", "status_code": 400}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Server error: {str(e)}", "status_code": 500}, status=500)
    return JsonResponse({"error": "Invalid HTTP method. Only POST allowed.", "status_code": 405}, status=405)
# @csrf_exempt
# def product_discount_inventory_view(request):
#     if request.method == 'POST':
#         try:
#             data = json.loads(request.body)
#             admin_id = data.get('admin_id')
#             action = data.get('action')
#             if not admin_id:
#                 return JsonResponse({"error": "Admin ID is required.", "status_code": 400}, status=400)
#             if not action or action not in ['inventory', 'discount']:
#                 return JsonResponse({"error": "Valid 'action' is required: 'inventory' or 'discount'.", "status_code": 400}, status=400)            
#             products = ProductsDetails.objects.filter(admin_id=admin_id).order_by('created_at')
#             if not products.exists():
#                 return JsonResponse({
#                     "message": "No products with discount found.",
#                     "status_code": 200,
#                     "admin_id": str(admin_id)
#                 }, status=200)
#             product_list = []
#             for product in products:
#                 if isinstance(product.product_images, list):
#                     image_urls = [
#                         f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{img}"
#                         for img in product.product_images
#                     ]
#                 elif isinstance(product.product_images, str):
#                     image_urls = [
#                         f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{product.product_images}"
#                     ]
#                 else:
#                     image_urls = []
#                 material_file_url = (
#                     f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{product.material_file}"
#                     if product.material_file else ""
#                 )
#                 price = float(product.price or 0)
#                 discount = float(product.discount or 0)
#                 gst = float(product.gst or 0)
#                 final_price = price - (price * discount / 100)
#                 product_data = {
#                     "product_id": str(product.id),
#                     "product_name": product.product_name,
#                     "sku_number": product.sku_number,
#                     "hsn_code":product.hsn_code,
#                     "price": round(price, 2),
#                     "gst": f"{round(gst, 2)}%",
#                     "final_price": round(final_price, 2),
#                     "discount": f"{round(discount)}%",
#                     "quantity": product.quantity,
#                     "material_file": material_file_url,
#                     "description": product.description,
#                     "number_of_specifications": product.number_of_specifications,
#                     "specifications": product.specifications,
#                     "product_images": image_urls,
#                     "created_at": product.created_at,
#                     "category": product.category.category_name if product.category else None,
#                     "sub_category": product.sub_category.sub_category_name if product.sub_category else None,
#                     "category_id": product.category_id,
#                     "sub_category_id": product.sub_category_id,
#                     "availability": product.availability,
#                     "product_status": product.product_status,
#                     "cart_status": product.cart_status,
#                 }
#                 if action == 'inventory':
#                     total_sold_quantity = OrderProducts.objects.filter(
#                         product_id=product.id,
#                         order_status='Paid'
#                     ).aggregate(total=Sum('quantity'))['total'] or 0
#                     product_data["total_quantity_sold"] = total_sold_quantity
#                 product_list.append(product_data)
#             return JsonResponse({
#                 "message": f"{action.capitalize()} products retrieved successfully.",
#                 "products": product_list,
#                 "status_code": 200,
#                 "admin_id": str(admin_id)
#             }, status=200)
#         except json.JSONDecodeError:
#             return JsonResponse({"error": "Invalid JSON data.", "status_code": 400}, status=400)
#         except Exception as e:
#             return JsonResponse({"error": f"An unexpected error occurred: {str(e)}", "status_code": 500}, status=500)
#     return JsonResponse({"error": "Invalid HTTP method. Only POST is allowed.", "status_code": 405}, status=405)


@csrf_exempt
def product_discount_inventory_view(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            admin_id = data.get('admin_id')
            action = data.get('action')
            if not admin_id:
                return JsonResponse({"error": "Admin ID is required.", "status_code": 400}, status=400)
            if not action or action not in ['inventory', 'discount']:
                return JsonResponse({"error": "Valid 'action' is required: 'inventory' or 'discount'.", "status_code": 400}, status=400)            
            products = ProductsDetails.objects.filter(
                admin_id=admin_id,
                category_id=data.get('category_id'),
                sub_category_id=data.get('sub_category_id')                                    
            ).order_by('created_at')
            if not products.exists():
                return JsonResponse({
                    "message": "No products with discount found.",
                    "status_code": 200,
                    "admin_id": str(admin_id)
                }, status=200)
            product_list = []
            for product in products:
                if isinstance(product.product_images, list):
                    image_urls = [
                        f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{img}"
                        for img in product.product_images
                    ]
                elif isinstance(product.product_images, str):
                    image_urls = [
                        f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{product.product_images}"
                    ]
                else:
                    image_urls = []
                material_file_url = (
                    f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{product.material_file}"
                    if product.material_file else ""
                )
                price = float(product.price or 0)
                discount = float(product.discount or 0)
                gst = float(product.gst or 0)
                final_price = price - (price * discount / 100)
                product_data = {
                    "product_id": str(product.id),
                    "product_name": product.product_name,
                    "sku_number": product.sku_number,
                    "hsn_code":product.hsn_code,
                    "price": round(price, 2),
                    "gst": f"{round(gst, 2)}%",
                    "final_price": round(final_price, 2),
                    "discount": f"{round(discount)}%",
                    "quantity": product.quantity,
                    "material_file": material_file_url,
                    "description": product.description,
                    "number_of_specifications": product.number_of_specifications,
                    "specifications": product.specifications,
                    "product_images": image_urls,
                    "created_at": product.created_at,
                    "category": product.category.category_name if product.category else None,
                    "sub_category": product.sub_category.sub_category_name if product.sub_category else None,
                    "category_id": product.category_id,
                    "sub_category_id": product.sub_category_id,
                    "availability": product.availability,
                    "product_status": product.product_status,
                    "cart_status": product.cart_status,
                }
                if action == 'inventory':
                    total_sold_quantity = OrderProducts.objects.filter(
                        product_id=product.id,
                        order_status='Paid'
                    ).aggregate(total=Sum('quantity'))['total'] or 0
                    product_data["total_quantity_sold"] = total_sold_quantity
                product_list.append(product_data)
            return JsonResponse({
                "message": f"{action.capitalize()} products retrieved successfully.",
                "products": product_list,
                "status_code": 200,
                "admin_id": str(admin_id)
            }, status=200)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON data.", "status_code": 400}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"An unexpected error occurred: {str(e)}", "status_code": 500}, status=500)
    return JsonResponse({"error": "Invalid HTTP method. Only POST is allowed.", "status_code": 405}, status=405)